"""Refresh the Portfolios dashboard from the live SharePoint workbook.

The workbook lives on the LWM Investment Team SharePoint site
('Copy of Model Portfolio Analysis - V4.3.11.xlsx'). This script reads it
through its OneDrive-synced copy on this machine, so whatever is on
SharePoint right now is what lands in the dashboard. It then rewrites the
seeded data block inside each dashboard HTML file listed in TARGETS.

Usage:
  python refresh_dashboard.py                       auto-locate the synced workbook
  python refresh_dashboard.py <path-to-xlsx>        explicit workbook path
  python refresh_dashboard.py --from-json <f.json>  debug: reuse an extracted full.json
"""
import collections
import datetime
import glob
import io
import json
import os
import re
import sys
import warnings

warnings.filterwarnings("ignore")

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
HOME = os.path.expanduser("~")

NOTES_PATH = os.path.join(REPO, "data", "notes_map.json")
FACTS_PATH = os.path.join(REPO, "data", "funds.json")
DOCS_PATH = os.path.join(REPO, "data", "documents_index.json")
RUN_LOG = os.path.join(REPO, "data", "run_log.json")

# Where the OneDrive client puts SharePoint content, for whichever user runs
# this. The first pattern is a synced team-site library, the second is the
# personal OneDrive (covers 'Add shortcut to OneDrive').
WORKBOOK_GLOBS = [
    os.path.join(HOME, "Lumin Wealth Management", "**", "*Model Portfolio Analysis*.xlsx"),
    os.path.join(HOME, "OneDrive - Lumin Wealth Management", "**", "*Model Portfolio Analysis*.xlsx"),
]


def targets():
    """Dashboard copies to update: the repo file, plus any synced SharePoint
    copy (e.g. a Dashboard folder the team syncs) found on this machine."""
    found = [os.path.join(REPO, "lumin-portfolios.html")]
    for root in ("Lumin Wealth Management", "OneDrive - Lumin Wealth Management"):
        pat = os.path.join(HOME, root, "**", "lumin-portfolios.html")
        for p in glob.glob(pat, recursive=True):
            if p not in found:
                found.append(p)
    return found

SYNC_HELP = """\
No synced copy of the workbook was found on this machine.

To link the dashboard to SharePoint, sync it once via OneDrive:
  1. Open the workbook's folder on the LWM Investment Team SharePoint site
     in your browser.
  2. Click 'Sync' (or 'Add shortcut to OneDrive') in the library toolbar.
  3. Wait for OneDrive to finish, then run this script again.

The synced file IS the SharePoint file - OneDrive keeps it current, and this
script always reads whatever SharePoint holds at refresh time.

Alternatively pass an explicit path:  python refresh_dashboard.py <path-to-xlsx>
"""

START = "  // ---------- seeded Lumin model data ----------"
END = "  // ---------- helpers ----------"

VALID_CLASSES = {"Equities", "Fixed Interest", "Diversifiers", "Cash"}
ISIN_RE = re.compile(r"^[A-Z]{2}[A-Z0-9]{9}\d$")

# Platform token -> (code, display name). Variants fold into their base platform.
PLATFORMS = {
    "WZ": ("WZ", "WealthZentrum"), "WZ2": ("WZ", "WealthZentrum"),
    "ASC": ("Asc", "Ascentric"), "ARC": ("ARC", "ARC"),
    "SL": ("SL", "Standard Life"), "AJB": ("AJB", "AJ Bell"),
    "RJ": ("RJ", "Raymond James"), "PRU": ("Pru", "Prudential"),
    "TRA": ("Tra", "Transact"), "ELEV": ("Elev", "Elevate"),
    "ELEV2": ("Elev", "Elevate"), "ELEVMG": ("ElevMG", "Elevate (M&G)"),
    "AXA": ("AXA", "AXA"), "UTM": ("UTM", "Utmost"),
    "SW": ("SW", "Scottish Widows"), "M&G": ("M&G", "M&G"),
    "FID": ("Fid", "Fidelity"), "FID SIPPS": ("Fid", "Fidelity"),
    "FIDSL SIPP": ("FidSL", "Fidelity SL SIPP"),
    "M&G OFFSHORE": ("M&GOff", "M&G Offshore"),
}

# Sheets whose names carry no year; resolved from the quarterly sequence around them.
DATE_OVERRIDE = {
    "Model Weights - From 30.10. (2)": "2025-10-30",
    "Model Weights - From 01.02. (2)": None,   # duplicate of 'Model Weights - From 01.02.2026'
}

# Columns that are flags, notes or client one-offs rather than models.
SKIP_COL = re.compile(r"^(column\d*|.*\?|\*note|jisa.*|freeman.*)$", re.I)

CODE = {"Equities": "E", "Fixed Interest": "F", "Diversifiers": "D", "Cash": "C"}
RANGES = ["Core", "Passive", "ESG", "Income"]

# Platforms are share-class variants of the same model. None seeds every
# platform's full history so the dashboard can filter by provider (decided
# 2026-08-12, reversing the 2026-08-11 ARC-only call). Set to a platform code
# (e.g. "ARC") to seed Core from that single platform and shrink the file.
CANONICAL_CORE_PLATFORM = None


def locate_workbook():
    hits = []
    for pat in WORKBOOK_GLOBS:
        for p in glob.glob(pat, recursive=True):
            base = os.path.basename(p)
            if base.startswith("~$"):
                continue
            hits.append(p)
    if not hits:
        return None
    # Prefer the file whose name matches SharePoint exactly (no '- Copy'
    # suffixes = not a detached duplicate), then the most recently saved.
    def rank(p):
        base = os.path.basename(p).lower()
        detached = base.count("- copy")
        return (detached, -os.path.getmtime(p))
    hits.sort(key=rank)
    return hits[0]


def sheet_date(name):
    if name in DATE_OVERRIDE:
        return DATE_OVERRIDE[name]
    m = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{2,4})", name)
    if not m:
        return None
    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if y < 100:
        y += 2000
    return "%04d-%02d-%02d" % (y, mo, d)


def parse_col(label):
    """'L50_Asc' / 'Fid30' / 'FIDSL SIPP L70' / 'M&G Offshore L50' -> (platform code, name, risk)."""
    s = label.strip()
    if SKIP_COL.match(s):
        return None

    m = re.match(r"^L(\d+)[_ ]+(.+)$", s, re.I)          # L50_Asc, L50_FID SIPPS
    if m:
        risk, plat = "L" + m.group(1), m.group(2)
    else:
        m = re.match(r"^(.+?)[_ ]*L(\d+)$", s, re.I)     # FIDSL SIPP L70, M&G Offshore L50
        if m:
            plat, risk = m.group(1), "L" + m.group(2)
        else:
            m = re.match(r"^([A-Za-z&]+?)(\d+)$", s)     # Fid30, SW50
            if not m:
                return None
            plat, risk = m.group(1), "L" + m.group(2)

    key = plat.strip().upper()
    if key not in PLATFORMS:
        return None
    code, name = PLATFORMS[key]
    return code, name, risk


def norm_class(raw):
    """Older sheets use ATS / Property / Government Bonds etc."""
    s = str(raw or "").strip().lower()
    if not s:
        return ""
    if s in ("equities", "equity"):
        return "Equities"
    if "fixed" in s or "bond" in s or "gilt" in s or "credit" in s:
        return "Fixed Interest"
    if s == "cash" or "money market" in s:
        return "Cash"
    if s in ("ats",) or "divers" in s or "propert" in s or "real estate" in s \
            or "infra" in s or "absolute" in s or "altern" in s or "commod" in s:
        return "Diversifiers"
    return ""


def clean(v):
    return str(v).strip() if v is not None else ""


# Subtotal and stat rows, plus bare asset-class labels used as section headers.
JUNK = {"placeholder", "ocf", "amc", "yield", "key facts", "equities",
        "equity", "fixed interest", "diversifiers", "ats", "property",
        "real estate", "government bonds", "total"}


def is_total(name):
    n = (name or "").strip().lower()
    if not n or n in JUNK:
        return True
    if n.startswith(("total", "non-cash", "#n/a", "#ref")):
        return True
    return bool(re.match(r"^[\d.,\-]+$", n))       # stray numeric rows


def extract(path):
    """Read the workbook and return (models list, warnings list)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # ---- master holdings: UNID / ISIN -> name, share class, asset class
    master, master_by_isin, master_by_name = {}, {}, {}
    for row in wb["Master Holdings List"].iter_rows(min_row=2, max_col=8, values_only=True):
        unid = row[2].strip() if isinstance(row[2], str) else ""
        if not unid:
            continue
        rec = {
            "name": str(row[3]).strip() if row[3] else "",
            "sc": str(row[4]).strip() if row[4] else "",
            "isin": str(row[5]).strip().upper() if row[5] else "",
            "asset": str(row[7]).strip() if row[7] else "",
        }
        master[unid] = rec
        if rec["isin"]:
            master_by_isin.setdefault(rec["isin"], rec)
        if rec["name"]:
            master_by_name.setdefault(rec["name"].strip().lower(), rec)

    def lookup(unid, isin, raw):
        return (master.get(unid) or master_by_isin.get(isin)
                or master_by_name.get((raw or "").strip().lower()) or {})

    def display_name(unid, isin, raw):
        rec = lookup(unid, isin, raw)
        name = rec.get("name") or raw or "(unnamed)"
        sc = rec.get("sc", "")
        if sc and sc.lower() not in name.lower():
            name += " (" + sc + ")"
        return name

    def asset_of(unid, isin, fallback, raw=""):
        rec = lookup(unid, isin, raw)
        a = norm_class(rec.get("asset", ""))
        return a or norm_class(fallback) or "Unclassified"

    def resolve_isin(unid, isin, raw):
        """Pre-2023 sheets carry no ISIN column - recover it from the master list."""
        if ISIN_RE.match(isin or ""):
            return isin
        rec = lookup(unid, "", raw)
        cand = rec.get("isin", "")
        return cand if ISIN_RE.match(cand or "") else ""

    # key (range, platform, risk) -> {meta, snapshots: {date: rows}}
    models = collections.OrderedDict()
    warns = []

    def put(rng, plat_code, plat_name, risk, date, rows, source):
        rows = [r for r in rows if r["weight"] > 0.00001]
        if not rows:
            return
        total = sum(r["weight"] for r in rows)
        # Kept either way - the dashboard flags an off-100% snapshot in the UI.
        if abs(total - 100) > 0.5:
            warns.append("%s %s %s @ %s totals %.2f%% (%s)" % (rng, plat_code, risk, date, total, source))
        key = (rng, plat_code, risk)
        if key not in models:
            models[key] = {"range": rng, "platform": plat_code, "platformName": plat_name,
                           "risk": risk, "snapshots": {}}
        if date in models[key]["snapshots"]:
            return
        models[key]["snapshots"][date] = {"rows": rows, "source": source}

    # ------------- Core: every dated 'Model Weights' sheet, every platform column
    sheets = [s for s in wb.sheetnames if s.lower().startswith("model weights")]

    # Gate: a 'Model Weights' sheet whose name doesn't parse to a date would be
    # silently dropped from history. Fail loudly instead - fix the sheet name
    # (Model Weights - From DD.MM.YYYY) or add it to DATE_OVERRIDE.
    unparseable = [s for s in sheets if sheet_date(s) is None and s not in DATE_OVERRIDE]
    if unparseable:
        raise SystemExit(
            "REFUSING TO RUN - these 'Model Weights' sheets have no parseable date "
            "and would be silently lost from the rebalance history:\n  "
            + "\n  ".join(unparseable)
            + "\nRename them 'Model Weights - From DD.MM.YYYY' (include the year) "
            "or add an entry to DATE_OVERRIDE in this script.")

    core_dates = []
    for sh in sheets:
        date = sheet_date(sh)
        if not date:
            continue
        core_dates.append(date)
        ws = wb[sh]
        grid = list(ws.iter_rows(max_col=80, values_only=True))
        if not grid:
            continue
        header = [clean(c) for c in grid[0]]

        def find(*names):
            for i, h in enumerate(header):
                if h.lower() in names:
                    return i
            return None

        ci_unid = find("unid")
        ci_name = find("holding", "holding name")
        ci_isin = find("isin")
        ci_asset = find("asset class")
        if ci_name is None:
            warns.append("skipped sheet '%s' - no Holding header" % sh)
            continue

        for ci, label in enumerate(header):
            if not label or ci in (ci_unid, ci_name, ci_isin, ci_asset):
                continue
            parsed = parse_col(label)
            if not parsed:
                continue
            code, pname, risk = parsed
            rows = []
            for r in grid[1:]:
                raw = clean(r[ci_name]) if ci_name < len(r) else ""
                if not raw or is_total(raw):
                    continue
                v = r[ci] if ci < len(r) else None
                if not isinstance(v, (int, float)):
                    continue
                unid = clean(r[ci_unid]) if ci_unid is not None and ci_unid < len(r) else ""
                rawisin = clean(r[ci_isin]).upper() if ci_isin is not None and ci_isin < len(r) else ""
                isin = resolve_isin(unid, rawisin, raw)
                fb = clean(r[ci_asset]) if ci_asset is not None and ci_asset < len(r) else ""
                rows.append({"fund": display_name(unid, isin, raw), "isin": isin,
                             "weight": round(float(v) * 100, 4),
                             "asset": asset_of(unid, isin, fb, raw)})
            put("Core", code, pname, risk, date, rows, sh)

    if not core_dates:
        raise SystemExit("no dated 'Model Weights' sheets found - is this the right workbook?")

    # Passive / ESG / Income sheets hold current allocations only; stamp them
    # with the latest rebalance date found in the Core history.
    current = max(core_dates)

    # ------------- Passive / ESG (WealthZentrum)
    def simple(sheet, cols, rng, headrow, name_i, isin_i, asset_i):
        if sheet not in wb.sheetnames:
            warns.append("sheet '%s' not found - %s range skipped" % (sheet, rng))
            return
        ws = wb[sheet]
        grid = list(ws.iter_rows(min_row=headrow, max_col=20, values_only=True))
        header = [clean(c) for c in grid[0]]
        for risk, want in cols:
            ci = None
            for i, h in enumerate(header):
                if h.lower() == want.lower():
                    ci = i
                    break
            if ci is None:
                warns.append("column '%s' not found in '%s'" % (want, sheet))
                continue
            rows = []
            for r in grid[1:]:
                raw = clean(r[name_i]) if name_i < len(r) else ""
                if not raw or is_total(raw):
                    continue
                v = r[ci] if ci < len(r) else None
                if not isinstance(v, (int, float)):
                    continue
                unid = clean(r[0])
                isin = resolve_isin(unid, clean(r[isin_i]).upper() if isin_i < len(r) else "", raw)
                fb = clean(r[asset_i]) if asset_i is not None and asset_i < len(r) else ""
                rows.append({"fund": display_name(unid, isin, raw), "isin": isin,
                             "weight": round(float(v) * 100, 4),
                             "asset": asset_of(unid, isin, fb, raw)})
            put(rng, "WZ", "WealthZentrum", risk, current, rows, sheet)

    simple("Passive Allocations", [("L30", "L30"), ("L50", "L50"), ("L70", "L70"), ("L100", "L100")],
           "Passive", 2, 1, 2, None)
    simple("ESG Allocations", [("L30", "L30 ESG"), ("L50", "L50 ESG"), ("L70", "L70 ESG"), ("L100", "L100 ESG")],
           "ESG", 1, 1, 2, 3)

    # ------------- Income (WealthZentrum)
    for risk, sheet in [("L30", "L30 Income"), ("L50", "L50 Income"),
                        ("L70", "L70 Income"), ("L100", "L100 Income")]:
        if sheet not in wb.sheetnames:
            warns.append("sheet '%s' not found - Income %s skipped" % (sheet, risk))
            continue
        ws = wb[sheet]
        grid = list(ws.iter_rows(min_row=3, max_col=7, values_only=True))
        header = [clean(c) for c in grid[0]]
        wi = header.index("Weight") if "Weight" in header else 5
        rows = []
        for r in grid[1:]:
            raw = clean(r[1])
            if not raw or is_total(raw):
                continue
            v = r[wi] if wi < len(r) else None
            if not isinstance(v, (int, float)):
                continue
            unid = clean(r[0])
            isin = resolve_isin(unid, clean(r[2]).upper(), raw)
            rows.append({"fund": display_name(unid, isin, raw), "isin": isin,
                         "weight": round(float(v) * 100, 4),
                         "asset": asset_of(unid, isin, "", raw)})
        put("Income", "WZ", "WealthZentrum", risk, current, rows, sheet)

    out = []
    for (rng, plat, risk), m in models.items():
        snaps = [{"date": d, "rows": v["rows"], "source": v["source"]}
                 for d, v in sorted(m["snapshots"].items(), reverse=True)]
        out.append({"range": rng, "platform": plat, "platformName": m["platformName"],
                    "risk": risk, "snapshots": snaps})
    return out, warns


def emit_block(models_in, notes_map, source_label, meta=None, facts=None, docs=None):
    """Compact-encode the models + manager notes into the SEED_* JS block.

    facts: the parsed data/funds.json (fund register - fees, yield, house,
    website per fund), seeded as SEED_FACTS so fund pages show workbook
    facts instead of hand-baked placeholders.
    docs: the parsed data/documents_index.json (scanned SharePoint
    document listings per dashboard section), seeded as SEED_DOCS.
    """
    # The platform list keeps every platform seen in the workbook (the UI offers
    # them all), even though Core models are seeded from the canonical one only.
    all_plats = sorted({(m["platform"], m["platformName"]) for m in models_in})

    if CANONICAL_CORE_PLATFORM:
        models_in = [m for m in models_in
                     if m["range"] != "Core" or m["platform"] == CANONICAL_CORE_PLATFORM]

    # Empty placeholder columns: a "model" that is only a cash line is not a model.
    dropped_empty = []
    for m in models_in:
        keep = []
        for s in m["snapshots"]:
            if sum(r["weight"] for r in s["rows"]) < 50:
                dropped_empty.append("%s %s %s @ %s" % (m["range"], m["platform"], m["risk"], s["date"]))
            else:
                keep.append(s)
        m["snapshots"] = keep
    models_in = [m for m in models_in if m["snapshots"]]

    funds, findex = [], {}
    dates, dindex = [], {}
    plats = [list(p) for p in all_plats]
    pindex = {p[0]: i for i, p in enumerate(all_plats)}

    def fund_id(r):
        key = (r["isin"] or r["fund"]).upper()
        if key not in findex:
            findex[key] = len(funds)
            funds.append([r["fund"], r["isin"], CODE.get(r["asset"], "U")])
        return findex[key]

    def date_id(d):
        if d not in dindex:
            dindex[d] = len(dates)
            dates.append(d)
        return dindex[d]

    def plat_id(code, name):
        if code not in pindex:
            pindex[code] = len(plats)
            plats.append([code, name])
        return pindex[code]

    out_models = []
    for m in sorted(models_in, key=lambda x: (RANGES.index(x["range"]) if x["range"] in RANGES else 9,
                                              x["platform"], int(re.sub(r"\D", "", x["risk"]) or 0))):
        snaps = []
        for s in sorted(m["snapshots"], key=lambda x: x["date"], reverse=True):
            holdings = []
            for r in sorted(s["rows"], key=lambda x: -x["weight"]):
                w = round(r["weight"], 4)
                holdings.append([fund_id(r), int(w) if w == int(w) else w])
            snaps.append([date_id(s["date"]), holdings])
        out_models.append([
            RANGES.index(m["range"]) if m["range"] in RANGES else 0,
            plat_id(m["platform"], m["platformName"]),
            m["risk"],
            snaps,
        ])

    # notes: fund index -> [label, path, file count]
    notes_out = {}
    for key, entries in notes_map.items():
        if key.upper() in findex:
            notes_out[findex[key.upper()]] = [[e["label"], e["path"], e["count"]] for e in entries]

    latest = max(dates) if dates else "unknown"

    def j(o):
        return json.dumps(o, ensure_ascii=False, separators=(",", ":"))

    b = io.StringIO()
    b.write(START + "\n")
    b.write("  // " + source_label + "\n")
    b.write("  // Weights are model allocations in %, largest first. No client data.\n")
    b.write("  // SEED_MODELS: [rangeIdx, platformIdx, risk, [[dateIdx, [[fundIdx, weight], ...]], ...]]\n\n")
    b.write("  var SEED_ASSET = { E: \"Equities\", F: \"Fixed Interest\", D: \"Diversifiers\", C: \"Cash\", U: \"Unclassified\" };\n")
    b.write("  var SEED_RANGES = " + j(RANGES) + ";\n")
    b.write("  var SEED_DATES = " + j(dates) + ";\n")
    b.write("  var SEED_PLATFORMS = " + j(plats) + ";\n\n")

    b.write("  var SEED_FUNDS = [\n")
    for f in funds:
        b.write("    " + j(f) + ",\n")
    b.write("  ];\n\n")

    b.write("  // Fund manager meeting notes on the shared drive, keyed by fund index.\n")
    b.write("  var SEED_NOTES = " + j(notes_out) + ";\n\n")

    # Fund register facts (from extract_funds.py via data/funds.json):
    # one record per fund, share classes grouped. The UI looks funds up
    # by ISIN (per share class) or by name.
    facts = facts or {}
    b.write("  // Fund register facts from the workbook (data/funds.json). Decimal fractions.\n")
    b.write("  var SEED_FACTS = " + j(facts.get("funds", [])) + ";\n")
    b.write("  var SEED_FACTS_META = " + j({
        "generatedAt": facts.get("generatedAt", ""),
        "sourceSheet": facts.get("sourceSheet", ""),
        "effectiveDate": facts.get("effectiveDate", ""),
    }) + ";\n\n")

    # Scanned document listings (data/documents_index.json): dashboard
    # section id -> [{title, url, modified, year}], newest first.
    docs = docs or {}
    b.write("  // Scanned SharePoint document listings, keyed by section id.\n")
    b.write("  var SEED_DOCS = " + j(docs.get("sections", {})) + ";\n")
    b.write("  var SEED_DOCS_META = " + j({
        "scannedAt": docs.get("scannedAt", ""),
    }) + ";\n\n")

    b.write("  var SEED_MODELS = [\n")
    for m in out_models:
        b.write("    [" + j(m[0]) + "," + j(m[1]) + "," + j(m[2]) + ",[\n")
        for s in m[3]:
            b.write("      [" + j(s[0]) + "," + j(s[1]) + "],\n")
        b.write("    ]],\n")
    b.write("  ];\n")

    meta = dict(meta or {})
    meta["latest"] = latest
    b.write("\n  // Provenance shown in the freshness banner. Written by refresh_dashboard.py.\n")
    b.write("  var SEED_META = " + j(meta) + ";\n")

    b.write("""
  var SEED_VERSION = "v3-%s";

  // Seeded models are reference data: rebuilt from SEED_* on every load and never
  // written to localStorage, which could not hold them. Only user edits persist.
  function hydrateSeed() {
    return SEED_MODELS.map(function (m) {
      var rangeName = SEED_RANGES[m[0]];
      var plat = SEED_PLATFORMS[m[1]];
      var risk = m[2];
      var id = "seed-" + [rangeName, plat[0], risk].join("-").toLowerCase();
      return {
        id: id,
        seed: true,
        range: rangeName,
        platform: plat[0],
        platformName: plat[1],
        risk: risk,
        note: "",
        snapshots: m[3].map(function (s, i) {
          return {
            id: id + "-" + i,
            seed: true,
            date: SEED_DATES[s[0]],
            note: "",
            author: "",
            rows: s[1].map(function (h) {
              var f = SEED_FUNDS[h[0]];
              return {
                fund: f[0], isin: f[1], weight: h[1],
                asset: SEED_ASSET[f[2]] || "Unclassified",
                notes: SEED_NOTES[h[0]] || null
              };
            })
          };
        })
      };
    });
  }
""" % latest)

    stats = {
        "funds": len(funds), "dates": len(dates), "platforms": len(plats),
        "models": len(out_models),
        "snapshots": sum(len(m[3]) for m in out_models),
        "notes": len(notes_out), "dropped": dropped_empty, "latest": latest,
    }
    return b.getvalue(), stats


def splice(target, block):
    html = open(target, encoding="utf-8").read()
    i, k = html.find(START), html.find(END)
    if i < 0 or k < 0 or k <= i:
        raise SystemExit("anchors not found in " + target)
    html = html[:i] + block + "\n" + html[k:]
    open(target, "w", encoding="utf-8").write(html)
    return len(html)


def check_regression(stats, allow_shrink):
    """Compare against the last run: fewer models/snapshots/funds than before
    usually means a renamed sheet or broken column, not a real change."""
    if not os.path.exists(RUN_LOG):
        return []
    try:
        runs = json.load(open(RUN_LOG, encoding="utf-8"))
    except ValueError:
        return []
    if not runs:
        return []
    prev = runs[-1]
    problems = []
    for k in ("models", "snapshots", "funds"):
        if stats[k] < prev.get(k, 0):
            problems.append("%s fell from %d to %d since the last run (%s)"
                            % (k, prev.get(k, 0), stats[k], prev.get("refreshedAt", "?")))
    if problems and not allow_shrink:
        raise SystemExit(
            "REFUSING TO PUBLISH - the extract shrank vs the previous run:\n  "
            + "\n  ".join(problems)
            + "\nIf this shrink is intentional (e.g. a model was retired), "
            "re-run with --allow-shrink. Otherwise check the workbook for a "
            "renamed sheet or a broken column before publishing.")
    return problems


def log_run(stats, meta, warns):
    runs = []
    if os.path.exists(RUN_LOG):
        try:
            runs = json.load(open(RUN_LOG, encoding="utf-8"))
        except ValueError:
            runs = []
    entry = {k: stats[k] for k in ("funds", "dates", "platforms", "models", "snapshots", "notes")}
    entry.update({"latest": stats["latest"], "warnings": len(warns),
                  "droppedEmpty": len(stats["dropped"])})
    entry.update(meta)
    runs.append(entry)
    os.makedirs(os.path.dirname(RUN_LOG), exist_ok=True)
    with open(RUN_LOG, "w", encoding="utf-8") as f:
        json.dump(runs[-200:], f, indent=1)


def main():
    argv = sys.argv[1:]
    allow_shrink = "--allow-shrink" in argv
    argv = [a for a in argv if a != "--allow-shrink"]

    if argv and argv[0] == "--from-json":
        models, warns = json.load(open(argv[1], encoding="utf-8")), []
        source_label = "Extracted from cached full.json (debug run)."
        meta = {"workbook": os.path.basename(argv[1]), "savedAt": "",
                "refreshedAt": datetime.datetime.now().strftime("%Y-%m-%d %H:%M")}
    else:
        path = argv[0] if argv else locate_workbook()
        if not path or not os.path.exists(path):
            sys.exit(SYNC_HELP)
        mtime = datetime.datetime.fromtimestamp(os.path.getmtime(path))
        print("workbook: %s" % path)
        print("saved:    %s" % mtime.strftime("%Y-%m-%d %H:%M"))
        models, warns = extract(path)
        meta = {"workbook": os.path.basename(path),
                "savedAt": mtime.strftime("%Y-%m-%d %H:%M"),
                "refreshedAt": datetime.datetime.now().strftime("%Y-%m-%d %H:%M")}
        source_label = ("Extracted from '%s' (SharePoint via OneDrive sync), "
                        "workbook saved %s, refreshed %s." % (
                            meta["workbook"], meta["savedAt"], meta["refreshedAt"]))

    notes_map = json.load(open(NOTES_PATH, encoding="utf-8")) if os.path.exists(NOTES_PATH) else {}
    if not notes_map:
        print("note: %s missing - manager-notes links will be empty" % NOTES_PATH)

    facts = json.load(open(FACTS_PATH, encoding="utf-8")) if os.path.exists(FACTS_PATH) else {}
    if not facts:
        print("note: %s missing - fund pages will show no register facts. "
              "Run scripts/extract_funds.py." % FACTS_PATH)

    docs = json.load(open(DOCS_PATH, encoding="utf-8")) if os.path.exists(DOCS_PATH) else {}
    if not docs:
        print("note: %s missing - seeded document listings will be empty. "
              "Run scripts/scan_documents.py." % DOCS_PATH)

    block, stats = emit_block(models, notes_map, source_label, meta,
                              facts=facts, docs=docs)

    # A fund register cut from an older rebalance than the model data is
    # stale - fees/usage flags may be wrong. Soft warning, not a gate.
    if facts and facts.get("effectiveDate") and facts["effectiveDate"] != stats["latest"]:
        warns.append("funds.json is from rebalance %s but model data is %s - "
                     "re-run scripts/extract_funds.py"
                     % (facts["effectiveDate"], stats["latest"]))

    print("funds %(funds)d | dates %(dates)d | platforms %(platforms)d | "
          "models %(models)d | snapshots %(snapshots)d | notes links %(notes)d" % stats)
    print("latest rebalance: %s" % stats["latest"])
    if stats["dropped"]:
        print("dropped empty snapshots: %d" % len(stats["dropped"]))

    check_regression(stats, allow_shrink)

    updated = 0
    for t in targets():
        if not os.path.exists(t):
            print("skipped (missing): %s" % t)
            continue
        size = splice(t, block)
        updated += 1
        print("updated: %s (%.0f KB)" % (t, size / 1024))
    if not updated:
        sys.exit("no dashboard HTML files found to update")

    log_run(stats, meta, warns)

    if warns:
        print("\nWARNINGS (%d):" % len(warns))
        for w in warns[:25]:
            print("   -", w)
        if len(warns) > 25:
            print("   ... and %d more" % (len(warns) - 25))


if __name__ == "__main__":
    main()
