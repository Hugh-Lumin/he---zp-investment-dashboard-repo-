# Extract the current Lumin fund register from the Model Portfolio Analysis workbook.
# Output: data/funds.json - one record per fund (grouped share classes), with
# current-usage flags per range (Core / Passive / ESG / Income), fees, yield,
# contacts and a best-effort fund house website link.
#
# Usage: python scripts/extract_funds.py [path-to-workbook]

import json
import re
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl

DEFAULT_WB = (
    r"C:\Users\HughEdwards\OneDrive - Lumin Wealth Management"
    r"\Investment team work\Copy of Model Portfolio Analysis - V4.3.11 - Copy.xlsx"
)

CORE_SHEET = "Model Weights - From 01.05.26"
CORE_EFFECTIVE = "2026-05-02"

# Fund house -> public website. Links go to the house homepage / UK fund centre;
# per-fund tearsheets use the FT link built from the ISIN instead (always resolvable).
HOUSE_SITES = {
    "Vanguard": "https://www.vanguardinvestor.co.uk/",
    "Fidelity": "https://www.fidelity.co.uk/funds/",
    "L&G": "https://fundcentres.lgim.com/",
    "Legal & General": "https://fundcentres.lgim.com/",
    "iShares": "https://www.ishares.com/uk",
    "Xtrackers": "https://etf.dws.com/en-gb/",
    "Invesco": "https://www.invesco.com/uk/en/financial-products.html",
    "Dimensional": "https://www.dimensional.com/gb-en",
    "PineBridge": "https://www.pinebridge.com/en/investment-strategies",
    "Polar Capital": "https://www.polarcapital.co.uk/gb/professional/Our-Funds/",
    "Premier Miton": "https://www.premiermiton.com/funds/",
    "Artemis": "https://www.artemisfunds.com/en/gbr/adviser/funds",
    "Redwheel": "https://www.redwheel.com/uk/en/professional/funds/",
    "Montanaro": "https://montanaro.co.uk/funds/",
    "M&G": "https://www.mandg.com/investments/professional-investor/en-gb/funds",
    "Janus Henderson": "https://www.janushenderson.com/en-gb/adviser/funds/",
    "TrinityBridge": "https://www.trinitybridge.com/investment-solutions",
    "Royal London": "https://www.rlam.com/uk/intermediaries/our-funds/",
    "Aberdeen": "https://www.aberdeeninvestments.com/en-gb/investor/funds",
    "abrdn": "https://www.aberdeeninvestments.com/en-gb/investor/funds",
    "Morgan Stanley": "https://www.morganstanley.com/im/en-gb/intermediary-investor/funds-and-performance.html",
    "CG": "https://cgasset.com/funds/",
    "Ruffer": "https://www.ruffer.co.uk/en/funds",
    "LF Ruffer": "https://www.ruffer.co.uk/en/funds",
    "Cohen & Steers": "https://www.cohenandsteers.com/funds/",
    "Amundi": "https://www.amundi.co.uk/professional/product-explorer",
    "UBS": "https://www.ubs.com/uk/en/assetmanagement/funds.html",
    "BNP Paribas": "https://www.bnpparibas-am.com/en-gb/professional/fund-explorer/",
    "HSBC": "https://www.assetmanagement.hsbc.co.uk/en/intermediary/investment-expertise/etfs",
    "JPM": "https://am.jpmorgan.com/gb/en/asset-management/adv/products/fund-explorer/",
    "Schroder": "https://www.schroders.com/en-gb/uk/intermediary/funds-and-strategies/funds-in-focus/",
    "BNY Mellon": "https://www.bny.com/investments/uk/en/intermediary/funds.html",
    "Chikara": "https://chikarainvestments.com/funds/",
    "PIMCO": "https://www.pimco.co.uk/en-gb/investments/gis",
    "TM Gravis": "https://www.graviscapital.com/funds",
    "ARC TIME": "https://time-investments.com/funds",
    "Fundsmith": "https://www.fundsmith.co.uk/",
    "Man GLG": "https://www.man.com/funds",
    "Jupiter": "https://www.jupiteram.com/uk/en/professional/funds/",
    "Liontrust": "https://www.liontrust.co.uk/what-we-offer/funds",
}


def norm(s):
    return str(s or "").strip()


def house_for(name):
    for house in sorted(HOUSE_SITES, key=len, reverse=True):
        if name.lower().startswith(house.lower()):
            return house, HOUSE_SITES[house]
    return name.split(" ")[0], None


def is_real_isin(v):
    return bool(re.fullmatch(r"[A-Z]{2}[A-Z0-9]{9}\d", norm(v).upper()))


def to_date(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return norm(v) or None


PASSIVE_PAT = re.compile(
    r"\b(index|tracker|etf|ftse|s&p|msci|500|all[- ]?share|all[- ]?world|epra|nareit|ultrashort)\b",
    re.I,
)

REGIONS = ["US", "UK", "Europe", "Japan", "Asia & EM"]

NAME_REGION_PATS = [
    ("UK", r"\b(uk|u\.k\.|united kingdom|gilts?|sterling|ftse (100|all[- ]share)|index[- ]linked)\b"),
    ("US", r"\b(us|u\.s\.|usa|united states|s&p|500|america[n]?)\b"),
    ("Europe", r"\b(europe|european)\b"),
    ("Japan", r"\b(japan|japanese)\b"),
    ("Asia & EM", r"\b(asia[n]?|emerging|em debt|emd|china|india)\b"),
]


def region_from_name(name):
    # Strip dots so "U.K." / "U.S." match the word-boundary patterns
    clean = name.replace(".", "")
    if re.search(r"\b(global|world|all[- ]world|international)\b", clean, re.I):
        return "Global"
    for region, pat in NAME_REGION_PATS:
        if re.search(pat.replace(r"u\.k\.", "uk").replace(r"u\.s\.", "us"), clean, re.I):
            return region
    return "Global"


def region_from_exposure(vec):
    # vec: [US, UK, Europe, Japan, Asia/EM] fractions. A region only counts as
    # dominant at >= 80% - MSCI World is ~72% US, so anything below that is a
    # global mandate, not a US fund. Components > 1.5 are sheet scale errors.
    vec = [v if 0 <= v <= 1.5 else 0 for v in vec]
    total = sum(vec)
    if total <= 0:
        return None
    best = max(range(5), key=lambda i: vec[i])
    return REGIONS[best] if vec[best] / total >= 0.8 else "Global"


def main():
    wb_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_WB
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)

    # ---- 1. Master Holdings List: metadata per security (keyed by UNID and ISIN)
    ws = wb["Master Holdings List"]
    rows = ws.iter_rows(values_only=True)
    header = [norm(h) for h in next(rows)]
    idx = {h: i for i, h in enumerate(header)}

    master_by_unid, master_by_isin = {}, {}
    for r in rows:
        if not r or not r[idx["UNID"]]:
            continue
        rec = {
            "unid": norm(r[idx["UNID"]]),
            "name": norm(r[idx["Holding Name"]]),
            "shareClass": norm(r[idx["Share Class"]]),
            "isin": norm(r[idx["ISIN"]]).upper(),
            "assetClass": norm(r[idx["Asset Class"]]),
            "active": str(r[idx["Is_Active"]]) == "True",
            "yield": r[idx["Yield"]],
            "yieldDate": to_date(r[idx["Yield Last Updated"]]),
            "ocf": r[idx["OCF/TER"]],
            "ocfDate": to_date(r[idx["OCF/TER Last Updated"]]),
            "amc": r[idx["AMC"]],
            "platforms": [p for p, col in [("WZ", "WZ"), ("Ascentric", "Ascentric?"), ("ARC", "ARC?")]
                          if str(r[idx[col]]) == "True"],
            "notes": norm(r[idx["Notes"]]),
            "contactPhone": norm(r[idx["Contact Telephone"]]),
            "contactEmail": norm(r[idx["Contact Email"]]),
        }
        master_by_unid[rec["unid"]] = rec
        if is_real_isin(rec["isin"]):
            master_by_isin[rec["isin"]] = rec

    # ---- 2. Current usage: (unid|isin) -> set of ranges, + peak weight seen
    usage = {}   # key -> {"ranges": set, "maxWeight": float, "name": str, "isin": str, "assetClass": str, "mgmt": str|None}

    def touch(key, rng, weight, name, isin, asset_class="", mgmt=None):
        u = usage.setdefault(key, {"ranges": set(), "maxWeight": 0.0, "name": name,
                                   "isin": isin, "assetClass": asset_class, "mgmt": mgmt})
        u["ranges"].add(rng)
        u["maxWeight"] = max(u["maxWeight"], weight)
        if asset_class and not u["assetClass"]:
            u["assetClass"] = asset_class
        if mgmt and not u["mgmt"]:
            u["mgmt"] = mgmt

    def num(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0

    # Core: any weight > 0 in any model column
    ws = wb[CORE_SHEET]
    rows = ws.iter_rows(values_only=True)
    hdr = [norm(h) for h in next(rows)]
    for r in rows:
        name, isin, ac = norm(r[1]), norm(r[2]).upper(), norm(r[3])
        if not name or name.lower().startswith("total") or name == "CASH":
            continue
        w = max((num(v) for v in r[4:len(hdr)]), default=0.0)
        if w > 0 and is_real_isin(isin):
            touch(isin, "Core", w, name, isin, ac)

    # Passive: cols J-M are L30..L100 weights
    ws = wb["Passive Allocations"]
    for r in ws.iter_rows(min_row=3, values_only=True):
        name, isin = norm(r[1]), norm(r[2]).upper()
        if not name or name.lower().startswith("total"):
            continue
        w = max(num(r[9]), num(r[10]), num(r[11]), num(r[12]))
        if w > 0 and is_real_isin(isin):
            touch(isin, "Passive", w, name, isin, "", "Passive")

    # ESG: cols N-Q are L30..L100 ESG weights
    ws = wb["ESG Allocations"]
    for r in ws.iter_rows(min_row=2, values_only=True):
        name, isin, ac = norm(r[1]), norm(r[2]).upper(), norm(r[3])
        if not name or name.lower().startswith("total"):
            continue
        w = max(num(r[13]), num(r[14]), num(r[15]), num(r[16]))
        if w > 0 and is_real_isin(isin):
            touch(isin, "ESG", w, name, isin, ac)

    # Income: per-level sheets; col F current weight, col D Active/Passive tag
    for lvl in ["L30", "L50", "L70", "L100"]:
        ws = wb[f"{lvl} Income"]
        for r in ws.iter_rows(min_row=4, max_col=7, values_only=True):
            name, isin, tag = norm(r[1]), norm(r[2]).upper(), norm(r[3])
            if not name or name.lower().startswith("total"):
                continue
            w = max(num(r[5]), num(r[6]))
            if w > 0 and is_real_isin(isin):
                touch(isin, "Income", w, name, isin, "",
                      tag if tag in ("Active", "Passive") else None)

    # ---- 2b. Regional exposure per ISIN, from the sheets that carry it
    exposure = {}  # isin -> [US, UK, Europe, Japan, AsiaEM]

    def record_exposure(isin, vec):
        if is_real_isin(isin) and any(v > 0 for v in vec) and isin not in exposure:
            exposure[isin] = vec

    # MIKE USE THIS ONE: cols D-H = US, UK, Europe, Japan, Asia/EM
    for r in wb["MIKE USE THIS ONE"].iter_rows(min_row=2, max_col=8, values_only=True):
        record_exposure(norm(r[2]).upper(), [num(r[3]), num(r[4]), num(r[5]), num(r[6]), num(r[7])])
    # Passive Allocations: cols E-I
    for r in wb["Passive Allocations"].iter_rows(min_row=3, max_col=9, values_only=True):
        record_exposure(norm(r[2]).upper(), [num(r[4]), num(r[5]), num(r[6]), num(r[7]), num(r[8])])
    # ESG Allocations: cols I-M
    for r in wb["ESG Allocations"].iter_rows(min_row=2, max_col=13, values_only=True):
        record_exposure(norm(r[2]).upper(), [num(r[8]), num(r[9]), num(r[10]), num(r[11]), num(r[12])])
    # Income analysis sheets: cols G-K (US % .. Asia / EM %)
    for lvl in ["L30", "L50", "L70", "L100"]:
        try:
            ws_i = wb[f"{lvl} Income 2"]
        except KeyError:
            continue
        for r in ws_i.iter_rows(min_row=3, max_col=11, values_only=True):
            record_exposure(norm(r[2]).upper(), [num(r[6]), num(r[7]), num(r[8]), num(r[9]), num(r[10])])

    # ---- 3. Group share classes into funds by master Holding Name (fallback: sheet name)
    funds = {}
    for key, u in usage.items():
        m = master_by_isin.get(u["isin"])
        fund_name = (m["name"] if m else u["name"]).strip()
        f = funds.setdefault(fund_name, {
            "name": fund_name,
            "assetClass": "",
            "mgmt": u["mgmt"],
            "ranges": set(),
            "shareClasses": [],
            "yield": None, "yieldDate": None,
            "contactPhone": "", "contactEmail": "", "notes": "",
        })
        f["ranges"] |= u["ranges"]
        ac = (m["assetClass"] if m else u["assetClass"]) or ""
        if ac and not f["assetClass"]:
            f["assetClass"] = ac
        if u["mgmt"] and not f["mgmt"]:
            f["mgmt"] = u["mgmt"]
        sc = {
            "label": m["shareClass"] if m else "",
            "isin": u["isin"],
            "ocf": (m["ocf"] if m else None),
            "ocfDate": (m["ocfDate"] if m else None),
            "amc": (m["amc"] if m else None),
            "platforms": (m["platforms"] if m else []),
        }
        if all(c["isin"] != sc["isin"] for c in f["shareClasses"]):
            f["shareClasses"].append(sc)
        if m:
            if m["yield"] is not None and f["yield"] is None:
                f["yield"], f["yieldDate"] = m["yield"], m["yieldDate"]
            for fld in ("contactPhone", "contactEmail", "notes"):
                if m[fld] and not f[fld]:
                    f[fld] = m[fld]

    # ---- 4. Finalise: house/website, mgmt fallback heuristic, sort
    out = []
    for f in sorted(funds.values(), key=lambda x: x["name"].lower()):
        house, site = house_for(f["name"])
        if not f["mgmt"]:
            f["mgmt"] = "Passive" if PASSIVE_PAT.search(f["name"]) else "Active"
        region = None
        for sc in f["shareClasses"]:
            vec = exposure.get(sc["isin"])
            if vec:
                region = region_from_exposure(vec)
                break
        f["region"] = region or region_from_name(f["name"])
        out.append({
            **f,
            "ranges": sorted(f["ranges"]),
            "house": house,
            "website": site,
            "shareClasses": sorted(f["shareClasses"], key=lambda c: c["isin"]),
        })

    payload = {
        "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "sourceSheet": CORE_SHEET,
        "effectiveDate": CORE_EFFECTIVE,
        "fundCount": len(out),
        "funds": out,
    }
    dest = Path(__file__).resolve().parent.parent / "data" / "funds.json"
    dest.parent.mkdir(exist_ok=True)
    dest.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    n_sc = sum(len(f["shareClasses"]) for f in out)
    print(f"Wrote {dest} - {len(out)} funds, {n_sc} share classes")
    no_site = [f["name"] for f in out if not f["website"]]
    if no_site:
        print("No website mapping for:", "; ".join(no_site))


if __name__ == "__main__":
    main()
